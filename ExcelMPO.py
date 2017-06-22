__author__ = 'Dali'

import openpyxl, datetime, os, warnings
from random import randint

def getDate():
    isValid = False
    while not isValid:
        date = (raw_input('Please enter date: (mm/dd/yyyy) \n'))
        try:
             d = datetime.datetime.strptime(date, '%m/%d/%Y').date()
             isValid = True
        except ValueError:
             print "Incorrect format"
    return d

def makeInvoices(x,y,sheet):
    today = datetime.datetime.strptime(str(datetime.datetime.today().date()), '%Y-%m-%d')
    noNewFolder = True
    newFolder = str(today.date())
    yearRental = 199.99
    halfYearRental = 150.00
    myStartDate = x
    myEndDate = y
    myMailboxes = sheet
    invoiceNum = randint(0,99999)
    answer = checkAnswer()
    isValid = False
    if answer is True:
        while not isValid:
            doubleCheck = (raw_input('Are you sure? (y / n) \n'))
            if doubleCheck in "y" or doubleCheck in "n":
                isValid = True
            else:
                print('incorrect format....')
        if doubleCheck in "n":
            print('ending program.....')
            raw_input('Press enter to exit')
        else:
            print ('writing invoices....')
            txtFileDates = open('Last Dates Used.txt', 'a')
            txtFileDates.write("on {0} --- from {1} to {2}\n".format (str(today.strftime('%m/%d/%Y')),str(myStartDate.strftime('%m/%d/%Y')), str(myEndDate.strftime('%m/%d/%Y'))))
            myInvoices = openpyxl.load_workbook('invoice.xlsx')
            invoiceSheet = myInvoices.get_sheet_by_name('Invoice')
            if os.path.exists(newFolder):
                os.chdir(newFolder)
                noNewFolder = False
            for i in range (2, myMailboxes.max_row+1,1):
                expiringDates = myMailboxes.cell(row=i,column=3).value
                if expiringDates is not None:
                    expiringDates = expiringDates.date()
                    if expiringDates >= myStartDate and expiringDates <= myEndDate:
                        if not os.path.exists(newFolder) and noNewFolder is True:
                                os.makedirs(newFolder)
                                os.chdir(newFolder)
                                noNewFolder = False

                        invoiceSheet.cell(coordinate="E2").value = str(today.strftime('%m/%d/%Y'))
                        invoiceSheet.cell(coordinate="E3").value = int(invoiceNum)
                        invoiceNum = invoiceNum + 1
                        invoiceSheet.cell(coordinate="E7").value = myMailboxes.cell(row=i,column=2).value
                        invoiceSheet.cell(coordinate="E9").value = myMailboxes.cell(row=i, column=1).value
                        months = myMailboxes.cell(row=i,column=4).value
                        if months == 12 and months is not None:
                            monthsYear = '6 months'
                            yearRental = halfYearRental
                        else:
                            monthsYear = '1 year'

                        invoiceSheet.cell(coordinate="B14").value = "%s mailbox rental due for box %s" % (str(monthsYear),str(myMailboxes.cell(row=i, column=1).value))
                        invoiceSheet.cell(coordinate="E14").value = yearRental
                        myInvoices.save(filename = str(myMailboxes.cell(row=i,column=2).value)+'.xlsx')
        print('ending program....')
        raw_input('Press enter to exit')

    else:
        for i in range (2, myMailboxes.max_row+1,1):
            expiringDates = myMailboxes.cell(row=i,column=3).value
            if expiringDates is not None:
                expiringDates = expiringDates.date()
                if expiringDates >= myStartDate and expiringDates <= myEndDate:
                    print(myMailboxes.cell(row = i, column = 1).value)
                    print (myMailboxes.cell(row = i, column = 2).value)
                    myDate = myMailboxes.cell(row=i,column=3).value
                    print(str(myDate.strftime('%m/%d/%Y')))
                    print ('------------------\n')
        print('ending program....')
        raw_input('Press enter to exit')



def checkAnswer():
    isValid = False
    while not isValid:
        answer = (raw_input('Would you like to make invoices for these mailboxes? (y or n) \n'))
        try:
            if answer in "y" or answer in "n":
                if answer in "n":
                    answer = False
                    isValid = True
                else:
                    answer = True
                    isValid = True
        except ValueError:
             print "Incorrect value"
    return answer





myStartDate = getDate()
myEndDate = getDate()
print ('opening workbook....')
warnings.filterwarnings("ignore")
wb = openpyxl.load_workbook('mailboxes.xlsx')
sheet = wb.get_sheet_by_name('NUMERIC LIST')




makeInvoices(myStartDate, myEndDate, sheet)

